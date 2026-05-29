"""Extract real stock movement data from Excel into JSON for the web app."""
import os
import sys
import openpyxl
import json
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import config

SRC = str(config.EXCEL_PATH)
OUT = str(config.JSON_PATH)

wb = openpyxl.load_workbook(SRC, data_only=True, read_only=True)

# Sheets that are summary pages (skip)
SKIP = {'สรุปสินค้าคงเหลือ', 'สรุปวัสดุคงเหลือ', 'Stock', 'รอรหัส'}

def clean_str(s: str) -> str:
    """Strip leading/trailing Thai standalone vowel marks and whitespace."""
    if s is None:
        return ''
    return str(s).strip().lstrip('ฺ์ัิีึืุูํ็่้๊๋').strip()


# Classify sheet category by name prefix
def classify(name: str) -> str:
    n = name.strip()
    # Numbered sheets "N.CODE" → finished-goods family
    if '.' in n:
        prefix, _, rest = n.partition('.')
        prefix = clean_str(prefix)
        code = clean_str(rest)
        if prefix and any(c.isdigit() for c in prefix):
            if code.startswith('BTA'):
                return 'BTA'
            return 'FG'
    if n.startswith('BOX'):
        return 'BOX'
    if n.startswith('PM') or n in ('Photo card', 'UMBRELLA', 'PMHANDKERCHIEF', 'PMMIRROR'):
        return 'PM'
    return 'OTHER'

def to_iso(dt):
    if isinstance(dt, datetime):
        return dt.strftime('%Y-%m-%d')
    return None

products = []

for sheet_name in wb.sheetnames:
    if sheet_name in SKIP:
        continue
    s = wb[sheet_name]

    # Read header info
    code = None
    name = None
    try:
        # Code at A6 or B10; name at B6
        rows_head = list(s.iter_rows(min_row=1, max_row=12, values_only=True))
        if len(rows_head) >= 6:
            r6 = rows_head[5]
            code = r6[0] if r6[0] else None
            name = r6[1] if len(r6) > 1 else None
        if not code and len(rows_head) >= 10:
            r10 = rows_head[9]
            code = r10[1] if len(r10) > 1 else None
    except Exception as e:
        print(f"Skip {sheet_name}: header read fail: {e}")
        continue

    if not code:
        # Use sheet name as fallback
        code = sheet_name.split('.', 1)[-1].strip() if '.' in sheet_name else sheet_name
    if not name:
        name = ''

    # Read transactions - need to scan full sheet to detect month headers
    transactions = []
    total_in = 0
    total_out = 0
    opening = 0
    closing = 0
    current_month = None  # detected from "Jan 2026" / "Feb 2026" headers
    seen_dates = set()

    MONTH_MAP = {'jan':1,'feb':2,'mar':3,'apr':4,'may':5,'jun':6,
                 'jul':7,'aug':8,'sep':9,'oct':10,'nov':11,'dec':12}

    for row in s.iter_rows(min_row=7, max_row=440, values_only=True):
        if not row or len(row) < 7:
            continue
        date_val, code_val, doc_no, brought, qty_in, qty_out, balance = row[0], row[1], row[2], row[3], row[4], row[5], row[6]
        note = row[7] if len(row) > 7 else None

        # Detect month header (e.g., "Jan 2026")
        if isinstance(date_val, str):
            low = date_val.strip().lower()
            if 'รวม' in date_val:
                continue
            for m_name, m_num in MONTH_MAP.items():
                if low.startswith(m_name):
                    current_month = m_num
                    break
            continue

        if date_val is None and brought is None and qty_in is None and qty_out is None:
            continue
        if not isinstance(date_val, datetime):
            continue

        # Fix date if month doesn't match header (Excel data entry error)
        if current_month and date_val.month != current_month:
            try:
                date_val = date_val.replace(month=current_month)
            except ValueError:
                # day out of range for target month (e.g. Feb 30)
                continue

        # Skip duplicates (defensive)
        date_key = date_val.date()
        if date_key in seen_dates:
            continue
        seen_dates.add(date_key)

        # Convert to numbers
        def num(x):
            if x is None or x == '':
                return 0
            try:
                return float(x)
            except (ValueError, TypeError):
                return 0

        b = num(brought)
        qi = num(qty_in)
        qo = num(qty_out)
        bal = num(balance)

        # First day opening
        if not transactions:
            opening = b

        total_in += qi
        total_out += qo
        closing = bal

        transactions.append({
            'date': to_iso(date_val),
            'doc_no': str(doc_no) if doc_no else '',
            'brought_forward': b,
            'in': qi,
            'out': qo,
            'balance': bal,
            'note': str(note) if note else ''
        })

    products.append({
        'sheet': sheet_name,
        'code': clean_str(code),
        'name': clean_str(name),
        'category': classify(sheet_name),
        'opening': opening,
        'total_in': total_in,
        'total_out': total_out,
        'closing': closing,
        'transactions': transactions,
    })
    print(f"{sheet_name}: {code} | {len(transactions)} rows | open={opening} in={total_in} out={total_out} close={closing}")

# Save
with open(OUT, 'w', encoding='utf-8') as f:
    json.dump({
        'generated_at': datetime.now().isoformat(),
        'source_file': SRC,
        'product_count': len(products),
        'products': products,
    }, f, ensure_ascii=False, indent=2)

print(f"\nDone. {len(products)} products -> {OUT}")
