"""Thorough verification: Excel vs PostgreSQL."""
import os
import sys
import openpyxl
import psycopg2
from datetime import datetime, date
from collections import defaultdict

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import config

DB = config.DB
SRC = str(config.EXCEL_PATH)
SKIP = {'สรุปสินค้าคงเหลือ', 'สรุปวัสดุคงเหลือ', 'Stock', 'รอรหัส'}

print("=" * 70)
print("VERIFICATION: Excel ⇄ PostgreSQL")
print("=" * 70)

# --- 1) Load DB state ---
conn = psycopg2.connect(**DB)
cur = conn.cursor()

cur.execute("SELECT id, sheet_name, code, opening_balance, total_in, total_out, closing_balance FROM stock_online.products")
db_products = {}  # sheet_name -> dict
for r in cur.fetchall():
    db_products[r[1]] = {
        'id': r[0], 'code': r[2],
        'opening': float(r[3]), 'in': float(r[4]),
        'out': float(r[5]), 'closing': float(r[6])
    }

cur.execute("""
    SELECT p.sheet_name, m.movement_date, m.qty_in, m.qty_out, m.balance
    FROM stock_online.stock_movements m
    JOIN stock_online.products p ON p.id = m.product_id
""")
db_movements = defaultdict(dict)  # sheet -> {date: (in, out, balance)}
for sheet, dt, qin, qout, bal in cur.fetchall():
    db_movements[sheet][dt] = (float(qin), float(qout), float(bal))

print(f"\n[DB] products={len(db_products)}  movements={sum(len(v) for v in db_movements.values()):,}")

# --- 2) Re-scan Excel ---
wb = openpyxl.load_workbook(SRC, data_only=True, read_only=True)

MONTH_MAP = {'jan':1,'feb':2,'mar':3,'apr':4,'may':5,'jun':6,
             'jul':7,'aug':8,'sep':9,'oct':10,'nov':11,'dec':12}

def extract_sheet(sheet_name):
    s = wb[sheet_name]
    rows_head = list(s.iter_rows(min_row=1, max_row=12, values_only=True))
    code = rows_head[5][0] if rows_head[5][0] else (rows_head[9][1] if rows_head[9][1] else sheet_name.split('.', 1)[-1].strip())

    transactions = {}
    total_in = total_out = opening = closing = 0
    current_month = None

    for row in s.iter_rows(min_row=7, max_row=440, values_only=True):
        if not row or len(row) < 7: continue
        date_val, _, _, brought, qty_in, qty_out, balance = row[0], row[1], row[2], row[3], row[4], row[5], row[6]
        if isinstance(date_val, str):
            if 'รวม' in date_val: continue
            low = date_val.strip().lower()
            for n, num in MONTH_MAP.items():
                if low.startswith(n):
                    current_month = num; break
            continue
        if date_val is None and brought is None and qty_in is None and qty_out is None: continue
        if not isinstance(date_val, datetime): continue
        if current_month and date_val.month != current_month:
            try: date_val = date_val.replace(month=current_month)
            except ValueError: continue
        d = date_val.date()
        if d in transactions: continue
        def n(x):
            try: return float(x) if x not in (None, '') else 0
            except: return 0
        b, qi, qo, bal = n(brought), n(qty_in), n(qty_out), n(balance)
        if not transactions: opening = b
        total_in += qi; total_out += qo; closing = bal
        transactions[d] = (qi, qo, bal)

    return {'code': str(code).strip(), 'opening': opening, 'in': total_in,
            'out': total_out, 'closing': closing, 'tx': transactions}

excel = {}
for sn in wb.sheetnames:
    if sn in SKIP: continue
    excel[sn] = extract_sheet(sn)

print(f"[XLSX] products={len(excel)}  movements={sum(len(v['tx']) for v in excel.values()):,}")

# --- 3) Compare product master ---
print("\n" + "=" * 70)
print("CHECK 1: Product master (code, opening, total_in, total_out, closing)")
print("=" * 70)
master_issues = 0
for sheet_name, xl in excel.items():
    if sheet_name not in db_products:
        print(f"  ✗ Missing in DB: {sheet_name}"); master_issues += 1; continue
    db = db_products[sheet_name]
    for fld in ('opening', 'in', 'out', 'closing'):
        if abs(xl[fld] - db[fld]) > 0.01:
            print(f"  ✗ {sheet_name}.{fld}: xlsx={xl[fld]} db={db[fld]}")
            master_issues += 1
    if xl['code'] != db['code']:
        print(f"  ✗ {sheet_name}.code: xlsx={xl['code']!r} db={db['code']!r}")
        master_issues += 1
for sheet_name in db_products:
    if sheet_name not in excel:
        print(f"  ✗ Extra in DB: {sheet_name}"); master_issues += 1

print(f"Master mismatches: {master_issues} / {len(excel) * 5} comparisons")

# --- 4) Compare movements (row-by-row) ---
print("\n" + "=" * 70)
print("CHECK 2: Daily movements (every date for every product)")
print("=" * 70)
total_compared = 0
mv_issues = 0
issue_samples = []
for sheet_name, xl in excel.items():
    if sheet_name not in db_movements: continue
    dbmv = db_movements[sheet_name]
    xlmv = xl['tx']
    # date set match
    xl_dates = set(xlmv.keys())
    db_dates = set(dbmv.keys())
    only_xl = xl_dates - db_dates
    only_db = db_dates - xl_dates
    if only_xl:
        mv_issues += len(only_xl)
        if len(issue_samples) < 5:
            issue_samples.append(f"  {sheet_name}: {len(only_xl)} dates only in Excel, e.g. {sorted(only_xl)[:3]}")
    if only_db:
        mv_issues += len(only_db)
        if len(issue_samples) < 5:
            issue_samples.append(f"  {sheet_name}: {len(only_db)} dates only in DB, e.g. {sorted(only_db)[:3]}")
    # value match
    for d in xl_dates & db_dates:
        total_compared += 1
        xi, xo, xb = xlmv[d]
        di, do, db_ = dbmv[d]
        if abs(xi - di) > 0.01 or abs(xo - do) > 0.01 or abs(xb - db_) > 0.01:
            mv_issues += 1
            if len(issue_samples) < 5:
                issue_samples.append(f"  {sheet_name} {d}: xlsx=(in={xi},out={xo},bal={xb}) db=(in={di},out={do},bal={db_})")

for s in issue_samples:
    print(s)
print(f"\nValues compared: {total_compared:,}")
print(f"Movement mismatches: {mv_issues}")

# --- 5) Spot check: random 5 products full diff ---
print("\n" + "=" * 70)
print("CHECK 3: Spot check - 5 sample products full summary")
print("=" * 70)
import random
random.seed(42)
samples = random.sample(list(excel.keys()), 5)
for sn in samples:
    xl = excel[sn]; db = db_products.get(sn, {})
    print(f"\n[{sn}]  code={xl['code']}")
    print(f"  Excel:  open={xl['opening']:>10}  in={xl['in']:>10}  out={xl['out']:>10}  close={xl['closing']:>10}  tx={len(xl['tx'])}")
    print(f"  DB:     open={db.get('opening',0):>10}  in={db.get('in',0):>10}  out={db.get('out',0):>10}  close={db.get('closing',0):>10}  tx={len(db_movements.get(sn,{}))}")
    # Count movement days where qty != 0
    xl_mov = sum(1 for v in xl['tx'].values() if v[0] or v[1])
    db_mov = sum(1 for v in db_movements.get(sn, {}).values() if v[0] or v[1])
    print(f"  Movement days (qty≠0):  Excel={xl_mov}  DB={db_mov}")

# --- 6) Final verdict ---
print("\n" + "=" * 70)
print("VERDICT")
print("=" * 70)
if master_issues == 0 and mv_issues == 0:
    print("✅ ข้อมูลตรงกัน 100% — Excel = PostgreSQL")
else:
    print(f"⚠️  Master issues: {master_issues}, Movement issues: {mv_issues}")

cur.close(); conn.close()
